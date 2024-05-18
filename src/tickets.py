import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def add_patient_record(file_path, status, diagnosis, name, email, chat_id, user_id, request, doctor):
    try:
        # Попробуем открыть существующую книгу
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        # Если файл не существует, создадим новую книгу
        workbook = Workbook()
        sheet = workbook.active
        # Добавим заголовки столбцов
        headers = ["Статус", "Диагноз", "Имя", "Почта", "chat_id", "user_id",  "Заявка", "Врач"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

    # Определим следующую строку для записи данных
    next_row = sheet.max_row + 1

    # Запишем данные пациента
    patient_data = [status, diagnosis, name, email, chat_id, user_id, request, doctor]
    for col_num, data in enumerate(patient_data, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}{next_row}"] = data

    # Сохраним изменения в файл
    workbook.save(file_path)
    #print(f"Запись для пациента {name} успешно добавлена в {file_path}")


TO_MAKE_TICKETS = ['вопрос онкологу','вопрос лимфологу', 'вопрос эндокринологу', 'вопрос диетологу',
     'вопрос дерматологу', 'хочу поделиться своим опытом', 'хочу оставить отзыв',
     'помочь по-другому']

def make_ticket(chat_id, user_id):
    pass
